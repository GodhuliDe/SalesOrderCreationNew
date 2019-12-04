'#===================================================================================================='
' Module:        ExcelDriver.py'
' Description:   Get and Set the values into the Data Sheet'
' Author:        Idris'
' Date Created:  28/11/2018'
'----------------------------------------------------------------------------------------------------'
' Modifications:'
' Date        Modified By      Change Description'
' ----------  ---------------  ----------------------------------------------------------------------'

'===================================================================================================='

import os
import pyodbc
objRootDir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
strRootDirectory = objRootDir.replace('\\','\\\\')
TestDataPath = strRootDirectory + "\\\\Resources\\\\TestData.xlsx"

from selenium.webdriver.common.by import By
from selenium import webdriver
from openpyxl import Workbook, load_workbook
RunManagerPath = strRootDirectory + "\\\\Resources\\\\RunManager.xlsx"
PutDataPath = strRootDirectory + "\\\\Resources\\\\Putdata.xlsx"
import pandas as pd

class ExcelDriver():
        
    
    def GetCellData(self,TestCaseName,columnName,SheetName):
        df = pd.read_excel(TestDataPath,SheetName)
        print("pathhh"+TestDataPath)
        print("TC"+TestCaseName)
        print((df.loc[df.TestCase==TestCaseName ,columnName])[0] )
        return (df.loc[df.TestCase==TestCaseName ,columnName])[0]   
        

    def PutCellData(self,TestCaseName,columnName,SheetName,cellValue):
       
        rowNum = ""
        colNum= ""
#         df = pd.read_excel(TestDataPath,SheetName)
#         (df.loc[df.TestCase==TestCaseName ,columnName])[0] = cellValue
        workbook = load_workbook(PutDataPath)
        sheet = workbook.sheet_by_name(SheetName)
        for row_cells in sheet.iter_rows(min_col=1, max_col=40):
            for cell in row_cells:
                if cell.value == columnName:
                    print(cell.value)
                    print(sheet.cell(rowNum,colNum+1).value)
                    ro = int (rowNum)
                    co = int(colNum+1)
                    (sheet.cell(row=ro,column=co).value)=cellValue
                    workbook.save(PutDataPath)
                    workbook.close()
                    break