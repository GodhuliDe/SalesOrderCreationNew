# CREATED BY Nagarjun K S on 20th November 2019
# Purpose: To automate the monitoring of Batch Job Status in Prod environment

#-----Import Libraries-----
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from cgitb import text
import openpyxl, pprint
import time
import xlsxwriter
import os
from tkinter import *
import tkinter as ttk

#-----Tkinter GUI start-----
root = Tk()
root.title("BatchJobs Automated")
startTime = StringVar()
endTime = StringVar()

#-----Excel sheet setup(one time process)-----
workbook = xlsxwriter.Workbook('ProdJobsdetails.xlsx')
worksheet = workbook.add_worksheet("batchjobs")
worksheet.set_column('A:A', 20)
worksheet.set_column('B:B', 20)
worksheet.set_column('C:C', 20)
worksheet.set_column('D:D', 20)

# .....beautify code for Column Header......#
HeaderStyle = workbook.add_format()
HeaderStyle.set_align("CENTER")
HeaderStyle.set_bold()
HeaderStyle.set_font_size(13)
HeaderStyle.set_bg_color('F51E1E')
HeaderStyle.set_font_color("FFFFFF")

# .....beautify code for Column Data elements......#
DataStyle = workbook.add_format()
DataStyle.set_align("CENTER")
DataStyle.set_bg_color('FEFDFD')
DataStyle.set_font_color("040000")

# .....Defining Names for Column Headers......#
worksheet.write("A2", "Batch Job", HeaderStyle)
worksheet.write("B2", "Status", HeaderStyle)
worksheet.write("C2", "Time", HeaderStyle)
worksheet.write("D2", "Comment", HeaderStyle)

# .....To iteratively write data values in excel from dictionary......#
def createReport(worksheet ,  index, tostore):
        print("Going to write the values")
        worksheet.write("A" + str(index + 2), str(tostore["Batch JOB"]), DataStyle)
        worksheet.write("B" + str(index + 2), str(tostore["Status"]), DataStyle)
        worksheet.write("C" + str(index + 2), str(tostore["Time"]), DataStyle)
        worksheet.write("D" + str(index + 2), str(tostore["Comment"]), DataStyle)

# .....to call Button to open Excel sheet......#
def openexcel():
    os.startfile("ProdJobsdetails.xlsx")

# .....On Clicking Start Button on Tkinter GUI the following block of code starts executing......#
def start():
    startT = startTime.get() + "/2019 12:00:00 AM"
    endT = endTime.get() + "/2019 11:59:59 PM"

    # .....Starting ChromeDriver and Navigating to the URL......#
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get('https://login.microsoftonline.com/rona.ca/wsfed?wa=wsignin1.0&wtrealm=spn%3a00000015-0000-0000-c000-000000000000&wctx=rm%3d0%26id%3dpassive%26ru%3d%252f%253ft%253d2019-11-06T14%25253a02%25253a09.7248130Z&wct=2019-11-06T14%3a02%3a20Z&wreply=https%3a%2f%2flcc-prd.operations.dynamics.com%2f')

    # .....SignIn Steps......#
    email = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH,'//input[@type="email"]')))
    email.send_keys("testatlasqa2@rona.ca")
    driver.find_element_by_xpath('//input[contains(@value, \'Next\')]').click()
    password = WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH,"//input[contains(@name, 'passwd')]")))
    password.send_keys("AtlasQA@123")
    signin = WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH,"//input[contains(@value, \'Sign in\')]")))
    signin.click()
    remember = WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH,"//input[contains(@value, \'Yes\')]")))
    remember.click()

    # .....For Navigating to BatchJobs steps......#
    sidepane = WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.XPATH,"//div[contains(@class,'modulesPane-opener')]")))
    sidepane.click()
    batchjobsnavigate = WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.XPATH,"//a[contains(text(),'Batch jobs')]")))
    batchjobsnavigate.click()

    # .....Using openyxl library to read BatchJob names form Excel sheet named Report.xlsx with sheet name ProdJobs......#
    wb = openpyxl.load_workbook('Report.xlsx')
    sheet = wb['ProdJobs']

    # .....Initializing parameters to be used in code ......#
    index = 1  # index -> iterating through different rows while writing into excel
    count = 0  #count -> to get count of number of jobs errored out
    slno = 1   #For assigning serial numbers for Batchjobs
    # .....Loop for reading different rows from Report.xlsx......#
    for row in range(2, sheet.max_row + 1):
        BatchJobs = [sheet['A' + str(row)].value]
        # .....Loop for Changing/Updating batchJobs in Job description menu of UI......#
        for job in BatchJobs:
            jobdescription = WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='columnHeader-label alignmentAuto'][contains(text(),'Job description')]")))
            jobdescription.click()
            jobname = WebDriverWait(driver,60).until(EC.visibility_of_element_located((By.XPATH,"//input[@name='FilterField_BatchJob_Caption_Caption_Input_0']")))
            jobname.clear()
            jobname.send_keys(job)
            print(str(slno) + ". " + job)
            slno = slno + 1
            driver.find_element_by_xpath("//span[@id='__BatchJob_Caption_ApplyFilters_1_label']").click()
            selectbatchjob = WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.XPATH,"//span[contains(text(),'Batch job')]")))
            selectbatchjob.click()

            # .....Navigatin to BatchJobs History of that particular Batch Job......#
            batchjobhistory = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"(//span[contains(text(),'Batch job history')])")))
            batchjobhistory.click()

            # .....For filtering by Timings......#
            actualstarttime = WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH,"(//div[contains(text(),'Actual start date/time')])[2]")))
            actualstarttime.click()
            arrowelement = WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH,"(//span[@class='button-label button-label-dropDown'])[9]")))
            arrowelement.click()
            driver.find_element_by_xpath("(//span[contains(text(),'between')])[1]").click()
            starttm = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH,"//input[@id='__FilterField_BatchJobHistory_StartDateTime_StartDateTime_Input_0_0_input']"))).send_keys(startT)
            time.sleep(0.5)
            endtm = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,"//input[@id='__FilterField_BatchJobHistory_StartDateTime_StartDateTime_Input_1_1_input']"))).send_keys(endT)
            driver.find_element_by_xpath("//span[@id='__BatchJobHistory_StartDateTime_ApplyFilters_3_label']").click()

            # .....For filtering the Status by Error......#
            statusmenu = WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.XPATH,"(//div[contains(text(),'Status')])[2]")))
            statusmenu.click()
            topassError = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH,"//input[@id='__FilterField_BatchJobHistory_Status_Status_Input_0_0_input']")))
            topassError.send_keys('Error')
            errorfilter = WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH,"//span[@id='__BatchJobHistory_Status_ApplyFilters_1_label']")))
            errorfilter.click()
            popuperrorMsg = ""
            errorMsg = ""
            time.sleep(2)

            #.....Exception handling to verify block of data is present after applying filters......#
            try:
                ele = driver.find_element_by_xpath("(//input[@class='textbox field displayoption viewMarker'])[5]")
                status = ele.get_attribute('title')
                if status == "Error":
                    print("Error")
                    errortim = driver.find_element_by_xpath("(//input[@ name='BatchJobHistory_StartDateTime'])[1]")
                    errortime = errortim.get_attribute('title')
                    print(errortime)
                    count = count + 1
                    log = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,"(//span[contains(text(),'Log')])[2]")))
                    log.click()
                    lightboxPresent = False
                    # .....Exception handling to verify Popup Error message......#
                    try:
                        lightbox = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "// div[ @class ='lightbox']")))
                        lightboxPresent = True
                    except:
                        lightboxPresent = False

                    # .....To get Error Message Details either from Popup or from Message details options under log......#
                    if lightboxPresent == True:
                        errorMsgTitle = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH,"//span[@id='titleField']")))
                        popuperrorMsg = errorMsgTitle.text
                        print(popuperrorMsg)
                        closeButton = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Close')]")))
                        closeButton.click()
                    else:
                        messageDetailsBtn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Message details')]")))
                        messageDetailsBtn.click()
                        WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "//div[contains(text(), 'Message details')]")))
                        messageDiv = driver.find_elements(By.XPATH, "//div[@class='prefixMessage-messageText']")
                        for i in range(0, len(messageDiv)- 1 ):
                            errorMsg = errorMsg + messageDiv[i].text + " | "
                        print("Message Details : " + errorMsg)
                        closeBtn = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH,"//button[@name='$Dialog-HideButton']")))
                        closeBtn.click()

                    # .....Creating a dictionary named 'tostore' for storing required key,values mentioned below......#
                    tostore = {}
                    tostore["Batch JOB"] = job
                    tostore["Status"] = status
                    tostore["Comment"] = popuperrorMsg + errorMsg
                    tostore["Time"] = errortime

                    # .....Calling createReport method to write values from dictionary to Excel sheet......#
                    createReport(worksheet, index, tostore)
                    index = index + 1
                else:
                    print("Ended")
            except Exception as e:
                print(e)
            time.sleep(0.5)
            driver.find_element_by_xpath("(//button[@name='SystemDefinedCloseButton'])[2]").click()

    # .....To write ErrorCount value to excel......#
    worksheet.write("A1", "Error Count = " + str(count), HeaderStyle)
    workbook.close()
    print("Error Count = " + str(count))
    driver.close()
    excelimage = ttk.PhotoImage(file="C:\\Users\\godde\\OneDrive - Lowe's Companies Inc\\Desktop\\SalesOrder\\SalesOrder\\TestPackage\\Scripts\\Excelpic.png")
    photoimage = excelimage.subsample(6, 6)
    excelbutton = ttk.Button(root, compound=ttk.TOP, width=50, height=50, image=photoimage, text="OPEN",
                             command=openexcel, font='Courier 8 bold')
    excelbutton.grid(column=1, row=2)
    excelbutton.image = photoimage

# .....Creating Labels,Entrybox and buttons from Tkinter GUI......#
ttk.Label(root, text="Start Date",font='Courier 12 bold').grid(column=0, row=0)
ttk.Entry(root, width=5, textvariable=startTime,font='Courier 12 bold').grid(column=1, row=0)
ttk.Label(root, text="End Date",font='Courier 12 bold').grid(column=0, row=1)
ttk.Entry(root, width=5, textvariable=endTime,font='Courier 12 bold').grid(column=1, row=1)
ttk.Button(root, text="Start",font='Courier 12 bold',fg='green', command=start).grid(column=0, row=2)
ttk.Label(root, text="Time should be in MM/DD format", fg='red',bg='lightgrey',font='Courier 10 bold').grid(column=0, row=3)
# ------------------------------------#
root.mainloop()
